import smtplib
import time
import openpyxl as openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By



def send_Mail(fromAdress,fromAdressPass,toAdress,message):
    try:
        mail = smtplib.SMTP("smtp.gmail.com", 587)
        mail.ehlo()
        mail.starttls()
        mail.login(fromAdress, fromAdressPass)
        mail.sendmail(fromAdress, toAdress,message)
        print("Mail has been sent")
        mail.close()
    except Exception as e:
        print("Mail could not be sent, Error:",e)

userNo="userNo"
password="password"
OBS_Firat=webdriver.Chrome("chromedriver.exe")
OBS_Firat.get("https://jasig.firat.edu.tr/cas/login?service=https://obs.firat.edu.tr/oibs/ogrenci")
OBS_Firat.find_element(by=By.NAME,value="username").send_keys(userNo)
OBS_Firat.find_element(by=By.NAME,value="password").send_keys(password)
OBS_Firat.find_element(by=By.CSS_SELECTOR,value="input[type=submit]").click()

while True:
    OBS_Firat.find_element(by=By.XPATH, value="//*[@id='proMenu']/li[3]").click()
    time.sleep(0.1)
    OBS_Firat.find_element(by=By.XPATH, value="//*[@id='proMenu']/li[3]/ul/li[2]").click()
    scoresURL=OBS_Firat.current_url
    OBS_Firat.switch_to.frame(OBS_Firat.find_element(by=By.ID,value="IFRAME1"))
    scoresHTML=OBS_Firat.page_source
    soup=BeautifulSoup(scoresHTML,'lxml')
    scoreList=soup.find("table",{"class":"grdStyle"})
    lessonList=[]
    lessonList.append(scoreList.find("span",{"id":"grd_not_listesi_lblSnv1_0"}).text.strip())
    lessonList.append(scoreList.find("span",{"id":"grd_not_listesi_lblSnv1_1"}).text.strip())
    lessonList.append(scoreList.find("span",{"id":"grd_not_listesi_lblSnv1_2"}).text.strip())
    lessonList.append(scoreList.find("span",{"id":"grd_not_listesi_lblSnv1_3"}).text.strip())
    lessonList.append(scoreList.find("span",{"id":"grd_not_listesi_lblSnv1_4"}).text.strip())
    lessonList.append(scoreList.find("span",{"id":"grd_not_listesi_lblSnv1_5"}).text.strip())


    readScores=openpyxl.load_workbook("Notlar.xlsx") 
    scoreTable=readScores.active


    for i in range(len(lessonList)):
        if lessonList[i].isdigit():
            scoreTable.cell(row=i+2, column=2).value=lessonList[i]
            readScores.save("Notlar.xlsx")

    for i in range(len(lessonList)):
        if scoreTable.cell(row=i+2, column=3).value != 1 and scoreTable.cell(row=i+2, column=2).value != -1:
            message = "Subject: Ders Notu\n\n{lessonName} notunuz:{lessonScore}".format(lessonName=scoreTable.cell(row=i+2, column=1).value, lessonScore=scoreTable.cell(row=i+2, column=2).value)
            print(message)
            send_Mail("fromAdress","fromAdressPass","toAdress",message)
            scoreTable.cell(row=i+2, column=3).value = 1
            readScores.save("Notlar.xlsx")


    OBS_Firat.get(scoresURL)
